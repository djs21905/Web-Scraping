import openpyxl
from openpyxl.utils import get_column_letter
from collections import Counter

'''
Written on 4-20-17
Function that takes a string argument as the filename of an excel spreadsheet.  
Before running the script make sure all values falling under the Log2 column are integers.
Please write all of your filenames in lowercase without spaces to prevent potential runtime errors. 
'''
def parse_excel(excel_file_name):

        # Opens the excel file using a workbook object
        wb = openpyxl.load_workbook(excel_file_name + '.xlsx', read_only = True, data_only = True)

        # selects the sheet 
        sheet = wb.get_sheet_by_name('significant 95%')

        # Retrieves the column Letters of both Log2 and GeneID
        log2_column = get_column_letter(sheet.max_column - 4)
        geneid_column = get_column_letter(sheet.max_column - 12)
        
        # Scans through the columns and appends each cell value to the list
        log2_list = [sheet[log2_column + str(row)].value for row in range(2,sheet.max_row + 1)]
        geneid_list = [sheet[geneid_column + str(row)].value for row in range(2,sheet.max_row + 1)]
        
        # Combines the GeneID and Log2 lists into a list of tuples 
        combined_list = list(zip(geneid_list, log2_list))

        # Keeps all tuples that are >1 or <-1
        final_list = [(i,excel_file_name) for i in combined_list if i[1]>1 or i[1]<-1]

        # A list containing tuples of the GeneID and Log2 value.  Also contains file identifier. 
        return final_list

        
def find_unique(list1,list2):

        list1 = list1 + list2
        final = []        
        while len(list1) > 0:
                popped = list1.pop()
                for index, gene in enumerate(list1):                        
                        if popped[0][0] == gene[0][0]:
                                del list1[index]
                                popped = ''
                                break                             
                final.append(popped)
                
                for index, gene in enumerate(final):
                        if gene == '':
                                del final[index] 
        return final


def write_to_excel(final,file_name):

        new_wb = openpyxl.Workbook()
        new_wb.get_sheet_names()
        sheet = new_wb.active

        rowNum = 1
        for item in final:
                print(item[0][0])
                sheet.cell(row = rowNum , column = 1).value = str(item[0][0])
                sheet.cell(row = rowNum , column = 2).value = float(item[0][1])
                sheet.cell(row = rowNum , column = 3).value = str(item[1]) 
                rowNum += 1
        new_wb.save(file_name +'.xlsx')

        
        

# Prompts the user to select two excel files
file_one = input('Type the wild type file name: ')
file_two = input('Type the mutant file name: ')


final_list = find_unique(parse_excel(file_one), parse_excel(file_two))
write_to_excel(final_list, file_one + file_two)



