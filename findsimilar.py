import openpyxl
from openpyxl.utils import get_column_letter


'''
Written on 4-27-17
Finds similar genes between fro7 and fro3 and Noo / Col

'''
def parse_excel(excel_file_name):
        
        # Opens the excel file using a workbook object
        wb = openpyxl.load_workbook(excel_file_name + '.xlsx', read_only = True, data_only = True)

        # selects the sheet 
        sheet = wb.get_sheet_by_name('Sheet')

        # Retrieves the column Letters of log2 geneID and identifier
        log2_column = get_column_letter(sheet.max_column - 1)
        geneid_column = get_column_letter(sheet.max_column - 2)
        identifier = get_column_letter(sheet.max_column)
 
        # Scans through the columns and appends each cell value to the list
        log2_list = [sheet[log2_column + str(row)].value for row in range(1,sheet.max_row + 1) if 'No' in sheet[identifier + str(row)].value or 'Col' in sheet[identifier + str(row)].value]
        geneid_list = [sheet[geneid_column + str(row)].value for row in range(1,sheet.max_row + 1) if'No' in sheet[identifier + str(row)].value or 'Col' in sheet[identifier + str(row)].value]
        identifier_list = [sheet[identifier + str(row)].value for row in range(1,sheet.max_row + 1) if 'No' in sheet[identifier + str(row)].value or 'Col' in sheet[identifier + str(row)].value]

        # Combines the GeneID and Log2 and Identifier lists into a list of tuples 
        final_list = list(zip(geneid_list, log2_list,identifier_list ))

        # A list containing tuples of the GeneID and Log2 value.  Also contains file identifier.
        return final_list

        
def find_similar(list1,list2):

        # Combine the two lists of Fro7 and Fro3
        combined_list = list1 + list2

  
        # Use Counter to tally
        similar_genes = []
       
        while len(combined_list) > 0:
                popped = combined_list.pop()
                for index, item in enumerate(combined_list):
                        if item[0] == popped[0]:
                                similar_genes.append(popped)
                                similar_genes.append(item)
                                del combined_list[index]
                        else:
                                pass
                popped = ''

        print(similar_genes)
        return similar_genes


def write_to_excel(final,file_name):

        new_wb = openpyxl.Workbook()
        new_wb.get_sheet_names()
        sheet = new_wb.active

        rowNum = 1
        for item in final:
                print(item[0])
                sheet.cell(row = rowNum , column = 1).value = str(item[0])
                sheet.cell(row = rowNum , column = 2).value = float(item[1])
                sheet.cell(row = rowNum , column = 3).value = str(item[2]) 
                rowNum += 1
        new_wb.save(file_name +'.xlsx')

        
        

# Prompts the user to select two excel files
file_one = input('Enter fro7 or Col data: ')
file_two = input('Enter fro3 or No-O data: ')

final_list = find_similar(parse_excel(file_one), parse_excel(file_two))
write_to_excel(final_list, file_one + file_two)



