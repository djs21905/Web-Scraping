import requests 
import openpyxl
from bs4 import BeautifulSoup 

'''
 -- Rutgers Oral History WebScraper 4/24/17 --
 This script functions to scrape data from the Rutgers Oral History DB
 http://oralhistory.rutgers.edu/alphabetical-index/161-z
 Names are compiled into a list.  The list names are broken down into sub-lists containing first
 last middle and suffix based on spaces between the names.  The data is then written to an .xlsx file

'''

url = input('Enter the url here: ') 
req = requests.get(url)
soup = BeautifulSoup(req.content)
table_data = soup.find_all('td')


raw_names = []
for data in table_data:
        names = data.find_all('a')
        for name in names:
                if len(name.text) > 5 and 'Rutgers' not in name.text:
                        raw_names.append(name.text)

seperated_names = [item.split(' ') for item in raw_names] 

# Write data to Excel file
wb = openpyxl.load_workbook('test.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

rowNum = sheet.max_row + 1 
for seperated_name in seperated_names:
        sheet.cell(row = rowNum, column = 1).value = rowNum - 1
        sheet.cell(row = rowNum, column = 2).value = seperated_name[0]
        if len(seperated_name) == 2:
                sheet.cell(row = rowNum, column = 3).value = seperated_name[1]
        if len(seperated_name) >= 3 and len(seperated_name)<4:
                sheet.cell(row = rowNum, column = 3).value = seperated_name[2]
                sheet.cell(row = rowNum, column = 4).value = seperated_name[1]
        elif len(seperated_name) >= 4:
                sheet.cell(row = rowNum, column = 5).value = seperated_name[3]
        rowNum += 1

wb.save('test.xlsx')

